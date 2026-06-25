"""
Excel template validation and tampering detection for AMS
"""
import os
import hashlib
import json
import zipfile
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
from .error_handlers import FileUploadError, ValidationError
from .logging_config import get_logger, security_logger

logger = get_logger(__name__)


class ExcelTemplateValidator:
    """Validates Excel files against approved templates"""
    
    # Approved template configurations
    APPROVED_TEMPLATES = {
        'faq_template': {
            'filename': 'faq_template.xlsx',
            'required_columns': [
                'Question', 'Answer', 'Category', 'State', 'Priority'
            ],
            'column_order': [
                'Question', 'Answer', 'Category', 'State', 'Priority'
            ],
            'required_data_fields': ['Question', 'Answer', 'State'],
            'max_rows': 1000,
            'allowed_extensions': ['xlsx'],
            'content_types': [
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ],
            'checksum': None,  # Will be set during initialization
            'metadata_signatures': [
                'AMS Template v1.0',
                'Audit Management System'
            ]
        }
    }
    
    def __init__(self, template_dir: str = None):
        self.template_dir = template_dir or os.path.join(
            os.path.dirname(os.path.dirname(__file__)), 'static'
        )
        self.template_checksums = {}
        self._initialize_templates()
    
    def _initialize_templates(self):
        """Initialize template checksums and metadata"""
        for template_name, config in self.APPROVED_TEMPLATES.items():
            template_path = os.path.join(self.template_dir, config['filename'])
            
            if os.path.exists(template_path):
                # Calculate checksum
                checksum = self._calculate_file_checksum(template_path)
                self.template_checksums[template_name] = checksum
                config['checksum'] = checksum
                
                logger.info(f"Template {template_name} initialized with checksum: {checksum}")
            else:
                logger.warning(f"Template file not found: {template_path}")
    
    def _calculate_file_checksum(self, file_path: str) -> str:
        """Calculate SHA-256 checksum of file"""
        sha256_hash = hashlib.sha256()
        
        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b""):
                sha256_hash.update(chunk)
        
        return sha256_hash.hexdigest()
    
    def _calculate_content_checksum(self, file_content: bytes) -> str:
        """Calculate SHA-256 checksum of file content"""
        return hashlib.sha256(file_content).hexdigest()
    
    def _extract_excel_metadata(self, file_path: str) -> Dict[str, Any]:
        """Extract metadata from Excel file"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            metadata = {
                'creator': workbook.properties.creator or '',
                'title': workbook.properties.title or '',
                'subject': workbook.properties.subject or '',
                'description': workbook.properties.description or '',
                'keywords': workbook.properties.keywords or '',
                'category': workbook.properties.category or '',
                'comments': workbook.properties.comments or '',
                'created': str(workbook.properties.created) if workbook.properties.created else '',
                'modified': str(workbook.properties.modified) if workbook.properties.modified else '',
                'application': workbook.properties.application or '',
                'app_version': workbook.properties.app_version or '',
                'sheets': [sheet.title for sheet in workbook.worksheets],
                'sheet_count': len(workbook.worksheets)
            }
            workbook.close()
            return metadata
        except Exception as e:
            logger.error(f"Failed to extract Excel metadata: {e}")
            return {}
    
    def _validate_file_structure(self, file_path: str, template_config: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Validate Excel file structure against template"""
        errors = []
        
        try:
            # Load workbook
            workbook = load_workbook(file_path, read_only=True)
            
            # Check if workbook has sheets
            if not workbook.worksheets:
                errors.append("Excel file has no worksheets")
                return False, errors
            
            # Use first worksheet
            worksheet = workbook.active
            
            # Get column headers from first row
            headers = []
            for cell in worksheet[1]:
                if cell.value is not None:
                    headers.append(str(cell.value).strip())
            
            # Check required columns
            required_columns = template_config['required_columns']
            missing_columns = set(required_columns) - set(headers)
            if missing_columns:
                errors.append(f"Missing required columns: {', '.join(missing_columns)}")
            
            # Check column order
            expected_order = template_config['column_order']
            if len(headers) >= len(expected_order):
                actual_order = headers[:len(expected_order)]
                if actual_order != expected_order:
                    errors.append(f"Column order mismatch. Expected: {expected_order}, Found: {actual_order}")
            
            # Check for extra columns
            allowed_columns = set(expected_order)
            extra_columns = set(headers) - allowed_columns
            if extra_columns:
                errors.append(f"Extra columns not allowed: {', '.join(extra_columns)}")
            
            # Check data rows
            max_rows = template_config.get('max_rows', 1000)
            data_rows = list(worksheet.iter_rows(min_row=2))  # Skip header row
            
            if len(data_rows) > max_rows:
                errors.append(f"Too many rows. Maximum allowed: {max_rows}, Found: {len(data_rows)}")
            
            # Validate required data fields
            required_data_fields = template_config['required_data_fields']
            for row_idx, row in enumerate(data_rows, start=2):
                row_data = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers):
                        row_data[headers[col_idx]] = cell.value
                
                # Check required fields in each row
                for field in required_data_fields:
                    if field in row_data and (row_data[field] is None or str(row_data[field]).strip() == ''):
                        errors.append(f"Row {row_idx}: Required field '{field}' is empty")
            
            workbook.close()
            
        except Exception as e:
            errors.append(f"Failed to validate Excel structure: {str(e)}")
        
        return len(errors) == 0, errors
    
    def _detect_tampering(self, file_path: str, template_config: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Detect signs of file tampering"""
        issues = []
        
        try:
            # Check file checksum
            expected_checksum = template_config.get('checksum')
            if expected_checksum:
                actual_checksum = self._calculate_file_checksum(file_path)
                if actual_checksum != expected_checksum:
                    issues.append("File checksum mismatch - possible tampering")
            
            # Check metadata signatures
            metadata = self._extract_excel_metadata(file_path)
            expected_signatures = template_config.get('metadata_signatures', [])
            
            signature_found = False
            for signature in expected_signatures:
                if signature.lower() in str(metadata).lower():
                    signature_found = True
                    break
            
            if not signature_found and expected_signatures:
                issues.append("Missing template signature - not an official template")
            
            # Check for suspicious content
            workbook = load_workbook(file_path, read_only=True)
            
            # Check for hidden sheets
            hidden_sheets = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == 'hidden']
            if hidden_sheets:
                issues.append(f"Hidden worksheets detected: {', '.join(hidden_sheets)}")
            
            # Check for very hidden sheets
            very_hidden_sheets = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == 'veryHidden']
            if very_hidden_sheets:
                issues.append(f"Very hidden worksheets detected: {', '.join(very_hidden_sheets)}")
            
            # Check for suspicious formulas or content
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # Formula
                            formula = str(cell.value).lower()
                            suspicious_keywords = ['shell', 'cmd', 'powershell', 'vbscript', 'javascript']
                            if any(keyword in formula for keyword in suspicious_keywords):
                                issues.append(f"Suspicious formula detected in sheet '{sheet.title}': {cell.value}")
            
            workbook.close()
            
            # Check file structure integrity
            try:
                with zipfile.ZipFile(file_path, 'r') as zip_file:
                    # Check for suspicious files in archive
                    suspicious_files = []
                    for file_info in zip_file.filelist:
                        filename = file_info.filename.lower()
                        if filename.startswith('xl/') and filename.endswith('.xml'):
                            # Check XML content for suspicious patterns
                            try:
                                with zip_file.open(file_info) as xml_file:
                                    content = xml_file.read().decode('utf-8', errors='ignore')
                                    if 'macro' in content.lower() or 'vba' in content.lower():
                                        suspicious_files.append(filename)
                            except (IOError, UnicodeDecodeError):
                                pass
                    
                    if suspicious_files:
                        issues.append(f"Suspicious content detected in: {', '.join(suspicious_files)}")
            
            except Exception as e:
                issues.append(f"Failed to check file integrity: {str(e)}")
            
        except Exception as e:
            issues.append(f"Tampering detection failed: {str(e)}")
        
        return len(issues) == 0, issues
    
    def validate_excel_file(self, file_path: str, template_name: str = 'faq_template') -> Dict[str, Any]:
        """Comprehensive Excel file validation"""
        
        # Check if template exists
        if template_name not in self.APPROVED_TEMPLATES:
            raise FileUploadError(f"Unknown template: {template_name}")
        
        template_config = self.APPROVED_TEMPLATES[template_name]
        
        validation_result = {
            'valid': False,
            'template_name': template_name,
            'errors': [],
            'warnings': [],
            'metadata': {},
            'checksum': None
        }
        
        try:
            # Check file extension
            file_ext = Path(file_path).suffix.lower().lstrip('.')
            if file_ext not in template_config['allowed_extensions']:
                validation_result['errors'].append(f"File extension '{file_ext}' not allowed")
                return validation_result
            
            # Extract metadata
            validation_result['metadata'] = self._extract_excel_metadata(file_path)
            
            # Calculate checksum
            validation_result['checksum'] = self._calculate_file_checksum(file_path)
            
            # Validate structure
            structure_valid, structure_errors = self._validate_file_structure(file_path, template_config)
            validation_result['errors'].extend(structure_errors)
            
            # Detect tampering
            tampering_ok, tampering_issues = self._detect_tampering(file_path, template_config)
            validation_result['errors'].extend(tampering_issues)
            
            # Overall validation result
            validation_result['valid'] = len(validation_result['errors']) == 0
            
            if validation_result['valid']:
                logger.info(f"Excel file validation passed: {template_name}")
            else:
                security_logger.log_suspicious_activity(
                    f"Excel validation failed for template {template_name}",
                    {
                        'errors': validation_result['errors'],
                        'checksum': validation_result['checksum'],
                        'metadata': validation_result['metadata']
                    }
                )
        
        except Exception as e:
            validation_result['errors'].append(f"Validation failed: {str(e)}")
            logger.error(f"Excel validation error: {e}")
        
        return validation_result
    
    def generate_template_signature(self, template_name: str) -> Dict[str, Any]:
        """Generate signature for approved template"""
        if template_name not in self.APPROVED_TEMPLATES:
            raise FileUploadError(f"Unknown template: {template_name}")
        
        template_config = self.APPROVED_TEMPLATES[template_name]
        template_path = os.path.join(self.template_dir, template_config['filename'])
        
        if not os.path.exists(template_path):
            raise FileUploadError(f"Template file not found: {template_path}")
        
        signature = {
            'template_name': template_name,
            'filename': template_config['filename'],
            'checksum': self._calculate_file_checksum(template_path),
            'metadata': self._extract_excel_metadata(template_path),
            'required_columns': template_config['required_columns'],
            'column_order': template_config['column_order'],
            'generated_at': datetime.utcnow().isoformat(),
            'version': '1.0'
        }
        
        return signature


class ExcelUploadValidator:
    """Main Excel upload validation service"""
    
    def __init__(self, template_dir: str = None):
        self.template_validator = ExcelTemplateValidator(template_dir)
    
    def validate_upload(self, file_storage, template_name: str = 'faq_template') -> Dict[str, Any]:
        """Validate uploaded Excel file"""
        
        # Save uploaded file temporarily
        filename = secure_filename(file_storage.filename)
        temp_path = os.path.join('/tmp', f"upload_{datetime.now().timestamp()}_{filename}")
        
        try:
            file_storage.save(temp_path)
            
            # Validate against template
            validation_result = self.template_validator.validate_excel_file(temp_path, template_name)
            
            # Add file information
            validation_result['original_filename'] = file_storage.filename
            validation_result['file_size'] = os.path.getsize(temp_path)
            validation_result['content_type'] = file_storage.content_type
            
            return validation_result
        
        finally:
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.remove(temp_path)
    
    def process_validated_excel(self, validation_result: Dict[str, Any]) -> pd.DataFrame:
        """Process validated Excel file into DataFrame"""
        if not validation_result['valid']:
            raise FileUploadError("Cannot process invalid Excel file")
        
        # This would be implemented based on specific business logic
        # For now, return a placeholder
        return pd.DataFrame()


# Global validator instance
excel_validator = ExcelUploadValidator()


def validate_excel_upload(file_storage, template_name: str = 'faq_template') -> Dict[str, Any]:
    """Validate Excel upload against approved template"""
    return excel_validator.validate_upload(file_storage, template_name)


def get_template_signature(template_name: str = 'faq_template') -> Dict[str, Any]:
    """Get template signature for validation"""
    return excel_validator.template_validator.generate_template_signature(template_name)
